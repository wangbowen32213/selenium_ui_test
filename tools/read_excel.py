import openpyxl
from config.conf import *
import sys
from tools.logger import log
sys.path.append(".")


def excel_to_list_header(excel_path,sheet_name):
    """
    读取有表头的excel文件，返回字典集合[{key:[value1,value2]}]
    :param excel_path:
    :param sheet_name:工作表名称
    :return:
    """
    log.debug(excel_path)
    log.debug(sheet_name)
    wb = openpyxl.load_workbook(excel_path)
    sheet = wb[sheet_name]
    row_lists = []
    header = [item.value for item in list(sheet.rows)[0]]
    for i in range(1,sheet.max_row):
        row_list = [item.value for item in list(sheet.rows)[i]]
        dict_row = dict(zip(header,row_list))
        row_lists.append(dict_row)
    return row_lists

def excel_to_list(excel_path,sheet_name):
    """
    读取无表头的excel文件，返回列表集合[[value1,value2],[]]
    :param excel_path:
    :param sheet_name:
    :return:
    """
    log.debug(excel_path)
    log.debug(sheet_name)
    wb = openpyxl.load_workbook(excel_path)
    sheet = wb[sheet_name]
    row_lists = []
    # header = [item.value for item in list(sheet.rows)[0]]
    for i in range(0,sheet.max_row):
        row_list = [item.value for item in list(sheet.rows)[i]]
        # dict_row = dict(zip(row_list))
        row_lists.append(row_list)
    return row_lists

def excel_to_list_row_li(excel_path,sheet_name="sheet1",row=None,li=None):
    """
    读取指定单元格的内容
    :param excel_path:
    :param sheet_name: 工作表名称
    :param row: 行数
    :param li: 列数
    :return:
    """
    log.debug(excel_path)
    log.debug(sheet_name)
    wb = openpyxl.load_workbook(excel_path)
    # sheet = wb[sheet_name]
    sheet = wb.active
    cell = sheet.cell(row,li)
    return cell.value

# def get_test_case(data_list,case_name):
#
#     for case_data in data_list:
#         if case_data["case_name"] == case_name:
#             log.debug(case_name)
#
#             return case_data
#         else:
#             log.debug("无命名为%s的用例" % case_name)
#             pass

if __name__ == '__main__':
    aa = excel_to_list_row_li("../package/test.xlsx","User_login",2,1)
    # cc = get_test_case(aa,"test_user_login_password_wrong")
    # print(cc["case_name"])
    print(aa)
